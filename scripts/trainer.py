# -*- coding:utf-8 -*-
from __future__ import print_function
import openpyxl
import re
import jieba
from model.frame_semantic_parser import *
from utils.conll_reader import *
from argparse import ArgumentParser
import codecs

class LabelSegment(object):
    def __init__(self, pos_expr, frame_expr):
        parts = pos_expr.split(',')
        self.start_pos = int(parts[0])
        self.end_pos = int(parts[1])
        self.frame_expr = frame_expr

    def __str__(self):
        return '{0},{1}<{2}>'.format(self.start_pos, self.end_pos, self.frame_expr)

    def __unicode__(self):
        return u'{0},{1}<{2}>'.format(self.start_pos, self.end_pos, self.frame_expr)

class TrainOperation(object):
    def __init__(self):
        pass

    def act(self):
        pass

class PosSample(TrainOperation):
    def __init__(self, items, model):
        super(PosSample, self).__init__()
        self.items = items
        self.model = model

    def act(self):
        loss, hit = self.model.pos_loss(self.items[0], self.items[1], self.items[3])
        return loss, hit, len(self.items[1]) - 1

class ConllSample(TrainOperation):
    def __init__(self, entries, model):
        super(ConllSample, self).__init__()
        self.entries = entries
        self.model = model

    def act(self):
        loss, hit = self.model.conll_loss(self.entries)
        return loss, hit, len(self.entries) - 1

class Trainer(object):
    SENTENCE_COLUMN = 1
    LABEL_COLUMN = 2
    POS_NEG_COLUMN = 3

    def __init__(self, corpus_path, penntree_bank_path, model_save_path, model):
        self.model = model
        # self.training_corpus = []
        self.training_samples = []
        if corpus_path:
            wrkbook = openpyxl.load_workbook(corpus_path)

            for tbl in wrkbook.worksheets:
                row_counts = tbl.max_row
                for i in xrange(1, 1 + row_counts):
                    sentence = tbl.cell(row = i, column = Trainer.SENTENCE_COLUMN).value
                    # sentence = sentence.strip()
                    label = tbl.cell(row = i, column = Trainer.LABEL_COLUMN).value
                    if label == None or sentence == None:
                        continue
                    labeled_segments = self.extract_frame_semantic(label)
                    frame_expr = u''.join([s.frame_expr for s in labeled_segments])
                    #looking for head idx in frame semantics
                    head = -1
                    for f in self.frame_semantics:
                        if f.expression(u'') == frame_expr:
                            head = f.head
                            break

                    assert head != -1
                    pos_neg = True if tbl.cell(row = i, column = Trainer.POS_NEG_COLUMN).value == 'POS' else False

                    #cut sentence, transfer sentence to token list
                    tokens = [w for w in jieba.cut(sentence)]
                    token_idx = self.range2token(tokens, labeled_segments)
                    if len(token_idx) == 0:
                        print(u'warning: {0} | {1}, token index not found'.format(sentence, label))
                    if token_idx[head - 1] > 0:
                        # self.training_corpus.append((tokens, token_idx, pos_neg, token_idx[head - 1]))
                        if pos_neg == True:
                            self.training_samples.append(PosSample((tokens, token_idx, pos_neg, token_idx[head - 1]), self.model))

        if penntree_bank_path:
            sentences = []
            with codecs.open(penntree_bank_path, encoding='utf-8', mode='r') as f:
                current_sentence = []
                for l in f:
                    parts = l.strip().split('\t')
                    id = int(parts[0])
                    token = parts[1]
                    parent_id = int(parts[2])
                    e = ConllEntry(id, token, None)
                    e.parent_id = parent_id
                    if id == 1:
                        sentences.append(current_sentence)
                        current_sentence = [e]
                    else:
                        current_sentence.append(e)
                    if len(parts) > 3:
                        e.fields['concept'] = parts[3]
                if len(current_sentence) > 0:
                    sentences.append(current_sentence)

            for e in sentences:
                self.training_samples.append(ConllSample(e, model))
        self.save_path = model_save_path
        self.trainer = AdamTrainer(self.model.model)

    def extract_frame_semantic(self, label):
        #find all content wrapped in <>, and then concatenate all of them
        pattern = re.compile('<(.*?)>')
        parts = pattern.split(label)
        total_pairs = len(parts) / 2
        pairs = [LabelSegment(parts[i * 2], parts[i * 2 + 1]) for i in xrange(total_pairs)]

        return pairs

    def range2token(self, tokens, ranges):
        token_range = {}
        idx = 0
        for t in tokens:
            token_range[t] = (idx, idx + len(t.token))
            idx += len(t.token)
        token_idx = []
        for r in ranges:
            found = False
            for i, t in enumerate(tokens):
                if token_range[t][0] == r.start_pos and token_range[t][1] == r.end_pos:
                    token_idx.append(i)
                    found = True
                    break
            if not found:
                token_idx.append(-1)
        return token_idx

    def train(self, epochs):
        eloss = 0
        hits = 0
        tokens = 0
        for turn in xrange(epochs):
            random.shuffle(self.training_samples)
            count = 0
            for s in self.training_samples:
                count += 1
                if count % 100 == 0:
                    print(u'proecessed: {2}, total loss: {0}, accuracy: {1}'.format(eloss, hits / max(1, float(tokens)), count))
                    hits = 0
                    tokens = 0
                    eloss = 0
                loss, hit, tken_count = s.act()
                hits += hit
                tokens += tken_count
                if loss != None:
                    eloss += loss.scalar_value()
                    loss.backward()
                    self.trainer.update()
                renew_cg()
            print('loop over: proecessed: {2}, total loss: {0}, accuracy: {1}'.format(eloss, hits / max(1, float(tokens)), count))
            hits = 0
            tokens = 0
            eloss = 0
            self.model.save_model(self.save_path.format(str(turn)))
            # for item in self.training_corpus:
            #     count += 1
            #     if count % 100 == 0:
            #         print(u'proecessed: {2}, total loss: {0}, accuracy: {1}'.format(eloss, hits / max(1, float(tokens)), count))
            #         hits = 0
            #         tokens = 0
            #         eloss = 0
            #     if len(item[1]) > 0:
            #         if item[2]:
            #             loss, hit = self.model.pos_loss(item[0], item[1], item[3])
            #             hits += hit
            #             tokens += len(item[1]) - 1
            #         else:
            #             loss = self.model.neg_loss(item[0], item[1], item[3])
            #         if loss != None:
            #             eloss += loss.scalar_value()
            #             loss.backward()
            #             self.trainer.update()
            #         renew_cg()

if __name__ == '__main__':
    #generate vocab_dict: word2idx
    jieba.load_userdict('../data/mydict')
    options = ArgumentParser()
    options.add_argument('--embedding_dim', action='store', dest='embedding_dim', type=int, default=100)
    options.add_argument('--char_embedding_dim', action='store', dest='char_embedding_dim', type=int, default = 30)
    options.add_argument('--lstm_hidden_dim', action='store', dest = 'lstm_hidden_dim', type = int, default=300)
    options.add_argument('--layers', action='store', dest='lstm_layers', type=int, default=3)
    options.add_argument('--attn_hidden', action='store', dest='attn_hidden', type=int, default=150)
    options.add_argument('--word_category_dim', action='store', dest='word_category_dim', type=int, default=30)
    options.add_argument('--vocab_capacity', action='store', dest='vocab_capacity', type=int, default=100000)
    options.add_argument('--concept_capacity', action='store', dest='concept_capacity', type=int, default=1000)
    options.add_argument('--char_capacity', action='store', dest='char_capacity', type=int, default=10000)
    options.add_argument('--postag_dim', action='store', dest='postag_dim', type=int, default=30)
    options.add_argument('--postag_label_size', action='store', dest='postag_label_size', type=int, default=20)
    options.add_argument('--token_conv_size', action = 'store', dest='token_conv_size', type = int, default=20)

    options = options.parse_args()
    char2idx = char_dict('../data/char')
    # model = FrameSemanticParser(concept_dict, word2idx, char2idx, options, '../data/semantics/frame_semantics.xlsx', matcher._model)
    model = FrameSemanticParser(char2idx, char2idx, options)
    word2idx = vocab_dict('../data/mydict')
    # model.load_model('../data/extractor.model')
    # model.load_model('../data/model/language10.model')
    # t = Trainer('../data/labeled_corpus.xlsx', '../data/semantics/frame_semantics.xlsx', '../data/lab/processed_conll.conll', '../data/model/lan2_{0}.model', matcher, model)
    t = Trainer(None, '../data/corpus/processed_conll.conll', '../data/model/ttttt{0}.model', model)

    t.train(1000)